#!/usr/bin/env python3
"""Ingest the CURE county-search source spreadsheet into the TitlePro docs tree.

Source workbook (4 sheets):
  - CA               : ~10 CA county recorder/tax rows + 2 test subjects each
  - FL               : ~67 FL county recorder/tax rows + 2 test subjects each
  - Sheet2           : ~614-row multi-state national inventory (search-scope, URLs)
  - FL Test Subjects : ~90-row richer per-order FL test-subject feed

Pipeline (deterministic + idempotent — re-running produces identical output):
  1. parse_workbook(xlsx) -> intermediate dict, also dumped to /tmp/sheet_ingest.json
  2. generate_* functions render/update the markdown + JSON deliverables from that dict

Sentinels normalized to None: "" / "x" / "X" / "NA" / "N/A" / "na" / "No Data Online"
/ "TBD" / "-" / "—". Real values are trimmed. Party-1/Party-2 columns are kept
distinct and joined with " + " for display.

Usage:
    PYTHONPATH=src python3 tools/ingest_county_search_sheet.py \
        --xlsx "/path/to/CA CURE County Search URLs (1).xlsx" [--json-only]
"""
from __future__ import annotations

import argparse
import json
import os
import re
from collections import Counter, OrderedDict
from datetime import date, datetime
from typing import Any, Dict, List, Optional

import openpyxl

# --------------------------------------------------------------------------- #
# Paths / constants
# --------------------------------------------------------------------------- #
REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
SYNCED = "2026-06-17"
DEFAULT_XLSX = "/Users/ag/Downloads/CA CURE County Search URLs (1).xlsx"
INTERMEDIATE_JSON = "/tmp/sheet_ingest.json"

# In-repo source-of-truth copy (dated). Header references point here.
IN_REPO_XLSX_REL = "docs/source_sheets/CA_CURE_County_Search_URLs_2026-06-17.xlsx"

FL_COUNTIES_DIR = os.path.join(REPO, "docs", "FL", "counties")
CA_COUNTIES_DIR = os.path.join(REPO, "docs", "CA", "counties")
NATIONAL_MD = os.path.join(REPO, "docs", "National_County_Inventory.md")
FL_TS_BATCH_MD = os.path.join(REPO, "docs", "FL", "FL_Test_Subjects_Batch_2026-06-17.md")
CA_EXAMPLES_MD = os.path.join(REPO, "docs", "CA_Examples.md")
FIXTURE_JSON = os.path.join(REPO, "tests", "fixtures", "test_subjects.json")

_SENTINELS = {
    "", "x", "na", "n/a", "no data online", "tbd", "-", "—", "none", "null",
}


# --------------------------------------------------------------------------- #
# Normalization helpers
# --------------------------------------------------------------------------- #
def norm(v: Any) -> Optional[str]:
    """Trim + collapse a cell; map sentinels to None. Never fabricate."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    s = re.sub(r"\s+", " ", s)
    if s.lower() in _SENTINELS:
        return None
    return s or None


def md_cell(v: Optional[str]) -> str:
    """Render a value for a markdown table cell. None -> em dash."""
    if v is None:
        return "—"
    return v.replace("|", "\\|").replace("\n", " ").strip()


def join_parties(p1: Optional[str], p2: Optional[str]) -> Optional[str]:
    parts = [p for p in (p1, p2) if p]
    return " + ".join(parts) if parts else None


def join_addr(addr: Optional[str], city: Optional[str]) -> Optional[str]:
    parts = [p for p in (addr, city) if p]
    return ", ".join(parts) if parts else None


def slug(name: str) -> str:
    s = name.strip().lower()
    s = s.replace("saint ", "st-").replace("st. ", "st-").replace(".", "")
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


# --------------------------------------------------------------------------- #
# Sheet parsers
# --------------------------------------------------------------------------- #
def parse_ca(ws) -> List[Dict[str, Any]]:
    out = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        county = norm(row[1])
        if not county:
            continue
        out.append(OrderedDict(
            date_researched=norm(row[0]),
            county=county,
            state=norm(row[2]) or "CA",
            recorder_url=norm(row[3]),
            recorder_captcha=norm(row[4]),
            post_captcha_url=norm(row[5]),
            exam_notes=norm(row[6]),
            tax_url=norm(row[7]),
            tax_captcha=norm(row[8]),
            tax_notes=norm(row[9]),
            subject_a=OrderedDict(
                party1=norm(row[10]), party2=norm(row[11]),
                address=norm(row[12]), city=norm(row[13]),
            ),
            subject_b=OrderedDict(
                party1=norm(row[15]), party2=norm(row[16]),
                address=norm(row[17]), city=norm(row[18]),
            ),
        ))
    return out


def parse_fl(ws) -> List[Dict[str, Any]]:
    out = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        county = norm(row[2])
        if not county:
            continue
        rank_raw = norm(row[0])
        try:
            rank = int(float(rank_raw)) if rank_raw is not None else None
        except (TypeError, ValueError):
            rank = None
        out.append(OrderedDict(
            population_rank=rank,
            status=norm(row[1]),
            county=county,
            state=norm(row[3]) or "FL",
            platform=norm(row[4]),
            recorder_apn_capable=norm(row[5]),
            doc_type_capable=norm(row[6]),
            auditor_apn_capable=norm(row[7]),
            recorder_url=norm(row[8]),
            recorder_captcha=norm(row[9]),
            post_captcha_url=norm(row[10]),
            exam_notes=norm(row[11]),
            tax_url=norm(row[12]),
            tax_platform=norm(row[13]),
            tax_captcha=norm(row[14]),
            tax_apn_capable=norm(row[15]),
            tax_post_captcha_url=norm(row[16]),
            tax_notes=norm(row[17]),
            subject_a=OrderedDict(
                party1=norm(row[18]), party2=norm(row[19]),
                address=norm(row[20]), city=norm(row[21]),
            ),
            subject_b=OrderedDict(
                party1=norm(row[23]), party2=norm(row[24]),
                address=norm(row[25]), city=norm(row[26]),
            ),
        ))
    return out


def parse_sheet2(ws) -> List[Dict[str, Any]]:
    """National inventory. col0=scope, col1=State, col2=County, col4=County URL,
    col6=Tax URL. Skip the stray header-leak row (state == 'State')."""
    out = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        scope = norm(row[0])
        state = norm(row[1])
        county = norm(row[2])
        county_url = norm(row[4])
        tax_url = norm(row[6])
        if not (state or county):
            continue
        # header-leak / product-name stray row
        if state == "State" or scope == "Product Name":
            continue
        out.append(OrderedDict(
            search_scope=scope,
            state=state,
            county=county,
            county_url=county_url,
            tax_url=tax_url,
        ))
    return out


def parse_fl_test_subjects(ws) -> List[Dict[str, Any]]:
    out = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        borrower = norm(row[5])
        if not borrower:
            continue
        street_no = norm(row[10])
        street_name = norm(row[11])
        addr = " ".join(p for p in (street_no, street_name) if p) or None
        out.append(OrderedDict(
            county_platform=norm(row[0]),
            document_type=norm(row[1]),
            date=norm(row[2]),
            vendor_reference_no=norm(row[3]),
            tag_reference_no=norm(row[4]),
            borrower=borrower,
            borrower_first_name=norm(row[6]),
            as_shown_on_document=norm(row[7]),
            street_no=street_no,
            street_name=street_name,
            address=addr,
            city=norm(row[12]),
            jurisdiction=norm(row[13]),
            state=norm(row[14]) or "FL",
        ))
    return out


def parse_workbook(xlsx: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    data = OrderedDict(
        _meta=OrderedDict(
            source_xlsx=xlsx,
            synced=SYNCED,
            sheets=list(wb.sheetnames),
        ),
        ca=parse_ca(wb["CA"]),
        fl=parse_fl(wb["FL"]),
        national=parse_sheet2(wb["Sheet2"]),
        fl_test_subjects=parse_fl_test_subjects(wb["FL Test Subjects"]),
    )
    return data


# --------------------------------------------------------------------------- #
# Markdown generators
# --------------------------------------------------------------------------- #
def _subject_rows(rec: Dict[str, Any]) -> List[List[str]]:
    rows = []
    for slot in ("a", "b"):
        s = rec[f"subject_{slot}"]
        party = join_parties(s["party1"], s["party2"])
        addr = join_addr(s["address"], s["city"])
        if party is None and addr is None:
            continue
        rows.append([slot.upper(), md_cell(party), md_cell(addr)])
    return rows


def render_fl_county_md(rec: Dict[str, Any]) -> str:
    rank = rec["population_rank"]
    status = rec["status"] or "Not yet reviewed"
    platform = rec["platform"] or "TBD"
    lines = []
    lines.append(f"# {rec['county']} County, FL — CURE Search Reference")
    lines.append("")
    lines.append(
        f"> **Population rank:** {rank if rank is not None else '—'}  |  "
        f"**Recorder platform:** {platform}  |  **Status:** {status}"
    )
    lines.append(
        f"> **Source of truth:** `{IN_REPO_XLSX_REL}` (FL sheet)  |  "
        f"**Synced:** {SYNCED}"
    )
    lines.append("")
    lines.append("## Recorder (Grantor/Grantee)")
    lines.append("")
    lines.append("| Field | Value |")
    lines.append("|---|---|")
    lines.append(f"| Search URL | {md_cell(rec['recorder_url'])} |")
    lines.append(f"| Recorder APN/PPN search capable | {md_cell(rec['recorder_apn_capable'])} |")
    lines.append(f"| Doc-type search capable | {md_cell(rec['doc_type_capable'])} |")
    lines.append(f"| Auditor site APN search capable | {md_cell(rec['auditor_apn_capable'])} |")
    lines.append(f"| CAPTCHA / disclaimer | {md_cell(rec['recorder_captcha'])} |")
    if rec["post_captcha_url"]:
        lines.append(f"| Post-CAPTCHA URL | {md_cell(rec['post_captcha_url'])} |")
    if rec["exam_notes"]:
        lines.append(f"| Exam notes | {md_cell(rec['exam_notes'])} |")
    lines.append("")
    lines.append("## Tax (Property Tax)")
    lines.append("")
    lines.append("| Field | Value |")
    lines.append("|---|---|")
    lines.append(f"| Tax URL | {md_cell(rec['tax_url'])} |")
    lines.append(f"| Tax platform | {md_cell(rec['tax_platform'])} |")
    lines.append(f"| CAPTCHA / disclaimer | {md_cell(rec['tax_captcha'])} |")
    lines.append(f"| APN search capable | {md_cell(rec['tax_apn_capable'])} |")
    if rec["tax_post_captcha_url"]:
        lines.append(f"| Post-CAPTCHA URL | {md_cell(rec['tax_post_captcha_url'])} |")
    if rec["tax_notes"]:
        lines.append(f"| Tax notes | {md_cell(rec['tax_notes'])} |")
    lines.append("")
    lines.append("## Test Subjects")
    lines.append("")
    lines.append("| Slot | Party | Address |")
    lines.append("|---|---|---|")
    srows = _subject_rows(rec)
    if srows:
        for r in srows:
            lines.append(f"| {r[0]} | {r[1]} | {r[2]} |")
    else:
        lines.append("| — | Not available per sheet | — |")
    lines.append("")
    return "\n".join(lines)


def render_ca_county_md(rec: Dict[str, Any], nn: int) -> str:
    lines = []
    lines.append(f"# {rec['county']} County, CA — CURE Search Reference")
    lines.append("")
    lines.append(
        f"> **Alpha index:** {nn:02d}  |  **Recorder platform:** TBD  |  "
        f"**Status:** ⏸️ CA paused (pivoted to FL 2026-05-20)"
    )
    lines.append(
        f"> **Source of truth:** `{IN_REPO_XLSX_REL}` (CA sheet)  |  "
        f"**Synced:** {SYNCED}"
    )
    if rec["date_researched"]:
        lines.append(f"> **Date researched:** {rec['date_researched']}")
    lines.append("")
    lines.append("## Recorder (Grantor/Grantee)")
    lines.append("")
    lines.append("| Field | Value |")
    lines.append("|---|---|")
    lines.append(f"| Search URL | {md_cell(rec['recorder_url'])} |")
    lines.append(f"| CAPTCHA / disclaimer | {md_cell(rec['recorder_captcha'])} |")
    if rec["post_captcha_url"]:
        lines.append(f"| Post-CAPTCHA URL | {md_cell(rec['post_captcha_url'])} |")
    if rec["exam_notes"]:
        lines.append(f"| Exam notes | {md_cell(rec['exam_notes'])} |")
    lines.append("")
    lines.append("## Tax (Property Tax)")
    lines.append("")
    lines.append("| Field | Value |")
    lines.append("|---|---|")
    lines.append(f"| Tax URL | {md_cell(rec['tax_url'])} |")
    lines.append(f"| CAPTCHA / disclaimer | {md_cell(rec['tax_captcha'])} |")
    if rec["tax_notes"]:
        lines.append(f"| Tax notes | {md_cell(rec['tax_notes'])} |")
    lines.append("")
    lines.append("## Test Subjects")
    lines.append("")
    lines.append("| Slot | Party | Address |")
    lines.append("|---|---|---|")
    srows = _subject_rows(rec)
    if srows:
        for r in srows:
            lines.append(f"| {r[0]} | {r[1]} | {r[2]} |")
    else:
        lines.append("| — | Not available per sheet | — |")
    lines.append("")
    return "\n".join(lines)


def generate_fl_counties(fl: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    os.makedirs(FL_COUNTIES_DIR, exist_ok=True)
    existing = {}
    for f in os.listdir(FL_COUNTIES_DIR):
        m = re.match(r"(\d+)_(.+)\.md$", f)
        if m:
            existing[m.group(2)] = f
    created, updated = [], []
    for rec in fl:
        sl = slug(rec["county"])
        rank = rec["population_rank"]
        if sl in existing:
            fname = existing[sl]
            updated.append(fname)
        else:
            nn = rank if rank is not None else 99
            fname = f"{nn:02d}_{sl}.md"
            created.append(fname)
        path = os.path.join(FL_COUNTIES_DIR, fname)
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(render_fl_county_md(rec))
    return {"created": created, "updated": updated}


def generate_ca_counties(ca: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    os.makedirs(CA_COUNTIES_DIR, exist_ok=True)
    created, updated = [], []
    # alphabetical NN
    ordered = sorted(ca, key=lambda r: r["county"].lower())
    for nn, rec in enumerate(ordered, start=1):
        sl = slug(rec["county"])
        fname = f"{nn:02d}_{sl}.md"
        path = os.path.join(CA_COUNTIES_DIR, fname)
        exists = os.path.exists(path)
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(render_ca_county_md(rec, nn))
        (updated if exists else created).append(fname)
    return {"created": created, "updated": updated}


def generate_national_inventory(national: List[Dict[str, Any]]) -> None:
    by_state: "OrderedDict[str, List[Dict[str, Any]]]" = OrderedDict()
    for rec in national:
        st = rec["state"] or "?"
        by_state.setdefault(st, []).append(rec)
    state_counts = {st: len(rows) for st, rows in by_state.items()}
    scope_counts = Counter(
        rec["search_scope"] for rec in national if rec["search_scope"]
    )
    total = len(national)

    n_with_url = sum(1 for r in national if r["county_url"] or r["tax_url"])

    lines = []
    lines.append("# National County Inventory — CURE Expansion Roadmap")
    lines.append("")
    lines.append(
        f"> **Source of truth:** `{IN_REPO_XLSX_REL}` (Sheet2)  |  "
        f"**Synced:** {SYNCED}"
    )
    lines.append(
        "> Multi-state county-URL inventory with per-county search-scope "
        "classification. This is the national expansion backlog beyond the "
        "CA + FL onboarded counties."
    )
    lines.append(
        "> **Source caveat:** Sheet2 has two URL header columns (`County URL`, "
        "`Tax URL`) but the `County URL` column is **entirely empty** in the "
        "source; only the `Tax URL` column carries data, and its contents are a "
        "mix of property-appraiser / GIS / recorder search portals (not strictly "
        "tax). The `URL` column below is that single captured URL, rendered "
        "verbatim — confirm the portal type during each county's live probe."
    )
    lines.append("")
    lines.append(
        f"**Total county rows:** {total}  |  **States:** {len(by_state)}  |  "
        f"**Rows with a URL:** {n_with_url}  |  **Rows missing URL:** "
        f"{total - n_with_url}"
    )
    lines.append("")
    lines.append("## Counts by State")
    lines.append("")
    lines.append("| State | Counties |")
    lines.append("|---|---|")
    for st in sorted(by_state):
        lines.append(f"| {st} | {state_counts[st]} |")
    lines.append("")
    lines.append("## Counts by Search-Scope Classification")
    lines.append("")
    lines.append("| # | Search Scope | Rows |")
    lines.append("|---|---|---|")
    for i, (scope, cnt) in enumerate(scope_counts.most_common(), start=1):
        lines.append(f"| {i} | {md_cell(scope)} | {cnt} |")
    n_no_scope = sum(1 for r in national if not r["search_scope"])
    if n_no_scope:
        lines.append(f"| — | *(no scope classification)* | {n_no_scope} |")
    lines.append("")
    lines.append("## Inventory by State")
    lines.append("")
    for st in sorted(by_state):
        rows = sorted(by_state[st], key=lambda r: (r["county"] or "").lower())
        lines.append(f"### {st} ({len(rows)})")
        lines.append("")
        lines.append("| Search Scope | County | URL (per source `Tax URL` col) |")
        lines.append("|---|---|---|")
        for r in rows:
            url = r["county_url"] or r["tax_url"]
            lines.append(
                f"| {md_cell(r['search_scope'])} | {md_cell(r['county'])} | "
                f"{md_cell(url)} |"
            )
        lines.append("")
    with open(NATIONAL_MD, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


def generate_fl_test_subjects_batch(ts: List[Dict[str, Any]]) -> None:
    by_juris: "OrderedDict[str, List[Dict[str, Any]]]" = OrderedDict()
    for rec in ts:
        # normalize jurisdiction casing for grouping (e.g. DUVAL/Duval)
        key = (rec["jurisdiction"] or "Unknown").strip()
        norm_key = key.title()
        by_juris.setdefault(norm_key, []).append(rec)

    lines = []
    lines.append("# FL Test Subjects — Batch 2026-06-17")
    lines.append("")
    lines.append(
        f"> **Source of truth:** `{IN_REPO_XLSX_REL}` (FL Test Subjects sheet)  |  "
        f"**Synced:** {SYNCED}"
    )
    lines.append(
        "> Richer per-order FL test-subject feed (vendor/TAG reference numbers, "
        "document type, recording date). ⚠️ Contains **real borrower names + "
        "addresses (PII)** — do not distribute outside the engagement."
    )
    lines.append("")
    lines.append(f"**Total subjects:** {len(ts)}  |  **Jurisdictions:** {len(by_juris)}")
    lines.append("")
    for juris in sorted(by_juris):
        rows = by_juris[juris]
        lines.append(f"## {juris} ({len(rows)})")
        lines.append("")
        lines.append("| Borrower | Doc Type | Date | Vendor Ref | TAG Ref | Address | City | Platform |")
        lines.append("|---|---|---|---|---|---|---|---|")
        for r in rows:
            borrower = r["borrower"]
            if r["borrower_first_name"]:
                borrower = f"{r['borrower_first_name']} {borrower}"
            if r["as_shown_on_document"]:
                borrower = f"{borrower} *(as shown: {r['as_shown_on_document']})*"
            lines.append(
                f"| {md_cell(borrower)} | {md_cell(r['document_type'])} | "
                f"{md_cell(r['date'])} | {md_cell(r['vendor_reference_no'])} | "
                f"{md_cell(r['tag_reference_no'])} | {md_cell(r['address'])} | "
                f"{md_cell(r['city'])} | {md_cell(r['county_platform'])} |"
            )
        lines.append("")
    with open(FL_TS_BATCH_MD, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


def merge_test_subjects_fixture(ts: List[Dict[str, Any]]) -> Dict[str, int]:
    """Append FL test subjects into tests/fixtures/test_subjects.json, preserving
    the existing {"subjects": {name: {...}}} dict shape and existing entries.

    Each FL subject is keyed by a stable display name and stored with the same
    top-level keys the fixture already uses (vested/liens/documents/party_mapping
    are empty placeholders — the source feed has no chain/lien data) plus a
    `meta` block carrying the order metadata. Re-running is idempotent: the
    ingest_2026_06_17 entries are rebuilt from scratch each run."""
    with open(FIXTURE_JSON, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    if "subjects" not in data or not isinstance(data["subjects"], dict):
        raise SystemExit("Unexpected fixture shape — aborting to avoid breakage")
    subjects = data["subjects"]

    # Drop any previously-ingested rows from this batch so re-runs stay clean.
    before_keys = set(subjects.keys())
    subjects = {
        k: v for k, v in subjects.items()
        if not (isinstance(v, dict) and v.get("meta", {}).get("ingest_batch") == SYNCED)
    }
    preserved = len(before_keys) - (len(before_keys) - len(subjects))

    added = 0
    seen = set()
    for r in ts:
        borrower = r["borrower"]
        if r["borrower_first_name"]:
            display = f"{r['borrower_first_name']} {borrower}".strip()
        else:
            display = borrower
        juris = (r["jurisdiction"] or "").title()
        key = f"{display} ({juris})" if juris else display
        # de-dup collisions by suffixing the TAG ref
        base_key = key
        if key in seen or key in subjects:
            suffix = r["tag_reference_no"] or r["vendor_reference_no"] or str(added)
            key = f"{base_key} [{suffix}]"
        seen.add(key)
        addr = join_addr(r["address"], r["city"])
        subjects[key] = {
            "vested": [],
            "liens": [],
            "documents": [],
            "party_mapping": {},
            "meta": {
                "ingest_batch": SYNCED,
                "source": "FL Test Subjects sheet",
                "borrower": borrower,
                "borrower_first_name": r["borrower_first_name"],
                "as_shown_on_document": r["as_shown_on_document"],
                "property_address": addr,
                "county": juris,
                "state": r["state"],
                "document_type": r["document_type"],
                "recorded_date": r["date"],
                "vendor_reference_no": r["vendor_reference_no"],
                "tag_reference_no": r["tag_reference_no"],
                "county_platform": r["county_platform"],
            },
        }
        added += 1

    data["subjects"] = subjects
    with open(FIXTURE_JSON, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)
        fh.write("\n")
    return {"added": added, "total": len(subjects)}


def refresh_ca_examples(ca: List[Dict[str, Any]]) -> None:
    """Refresh ONLY the '## Counties + Subjects' table in docs/CA_Examples.md,
    preserving the surrounding prose (header warning, already-run cases, pending
    cases). If the file doesn't exist, create a minimal version."""
    table_lines = []
    table_lines.append("| # | County | Subject A | Address A | Subject B | Address B |")
    table_lines.append("|---|---|---|---|---|---|")
    ordered = sorted(ca, key=lambda r: r["county"].lower())
    for i, rec in enumerate(ordered, start=1):
        a, b = rec["subject_a"], rec["subject_b"]
        sa = join_parties(a["party1"], a["party2"]) or "Not available per sheet"
        aa = join_addr(a["address"], a["city"]) or "—"
        sb = join_parties(b["party1"], b["party2"]) or "Not available per sheet"
        ab = join_addr(b["address"], b["city"]) or "—"
        table_lines.append(
            f"| {i} | **{rec['county']}** | {md_cell(sa)} | {md_cell(aa)} | "
            f"{md_cell(sb)} | {md_cell(ab)} |"
        )
    table_block = "\n".join(table_lines)

    if os.path.exists(CA_EXAMPLES_MD):
        with open(CA_EXAMPLES_MD, "r", encoding="utf-8") as fh:
            content = fh.read()
        # Replace the table that lives under '## Counties + Subjects' up to the
        # next '## ' heading. Keep everything else. Group 1 is just the heading
        # line; the replacement re-inserts exactly one blank line + table so the
        # spacing is normalized and re-runs are idempotent.
        pat = re.compile(
            r"(## Counties \+ Subjects\n).*?(\n## )",
            re.DOTALL,
        )
        if pat.search(content):
            new = pat.sub(
                lambda m: m.group(1) + "\n" + table_block + "\n" + m.group(2),
                content,
                count=1,
            )
            # bump the source-of-truth line if present
            new = re.sub(
                r"> \*\*Source:\*\*.*",
                f"> **Source:** `{IN_REPO_XLSX_REL}` (CA sheet, last sync **{SYNCED}**)",
                new,
                count=1,
            )
            content = new
        else:
            content = content.rstrip() + "\n\n## Counties + Subjects\n\n" + table_block + "\n"
        with open(CA_EXAMPLES_MD, "w", encoding="utf-8") as fh:
            fh.write(content)
    else:
        with open(CA_EXAMPLES_MD, "w", encoding="utf-8") as fh:
            fh.write(
                "# CA Examples — Test Subjects (Subject A + Subject B per county)\n\n"
                f"> **Source:** `{IN_REPO_XLSX_REL}` (CA sheet, last sync **{SYNCED}**)\n\n"
                "## Counties + Subjects\n\n" + table_block + "\n"
            )


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--json-only", action="store_true",
                    help="Only parse to /tmp/sheet_ingest.json; skip file generation")
    args = ap.parse_args()

    data = parse_workbook(args.xlsx)
    with open(INTERMEDIATE_JSON, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)
    print(f"[ingest] parsed -> {INTERMEDIATE_JSON}")
    print(f"  CA counties:        {len(data['ca'])}")
    print(f"  FL counties:        {len(data['fl'])}")
    print(f"  National rows:      {len(data['national'])}")
    print(f"  FL test subjects:   {len(data['fl_test_subjects'])}")

    if args.json_only:
        return

    fl_res = generate_fl_counties(data["fl"])
    ca_res = generate_ca_counties(data["ca"])
    generate_national_inventory(data["national"])
    generate_fl_test_subjects_batch(data["fl_test_subjects"])
    fixture_res = merge_test_subjects_fixture(data["fl_test_subjects"])
    refresh_ca_examples(data["ca"])

    print(f"[gen] FL counties:  {len(fl_res['updated'])} updated, "
          f"{len(fl_res['created'])} created")
    print(f"[gen] CA counties:  {len(ca_res['created'])} created, "
          f"{len(ca_res['updated'])} updated")
    print(f"[gen] National inventory -> {NATIONAL_MD}")
    print(f"[gen] FL test-subjects batch -> {FL_TS_BATCH_MD}")
    print(f"[gen] fixture: +{fixture_res['added']} subjects, "
          f"{fixture_res['total']} total")
    print(f"[gen] CA_Examples refreshed -> {CA_EXAMPLES_MD}")


if __name__ == "__main__":
    main()
